import requests
import openpyxl

# Define the API key and base URL for Ola Maps
API_KEY = "gqGVnKZkmeOGHl97NnhaJxtVq2PFVzezCdEPyx6S"
BASE_URL = "https://api.olamaps.io/places/v1/geocode"

# Read the Excel file
input_file = 'Goverment Data.xlsx'
output_file = 'Geocoded_Data.xlsx'

# Load the workbook and select the active sheet
print("Loading input Excel file...")
wb = openpyxl.load_workbook(input_file)
sheet = wb.active

# Create a new workbook for the results
wb_output = openpyxl.Workbook()
sheet_output = wb_output.active

# Write the header to the new Excel sheet
headers = ["Sr. No.", "Fitment Center Name", "Address", "Licence No.", "Latitude", "Longitude"]
sheet_output.append(headers)

# Loop through each row in the input file starting from the second row
print("Starting geocoding process...")
for row in sheet.iter_rows(min_row=2, values_only=True):
    sr_no, center_name, address, licence_no = row
    print(f"Processing row {sr_no}: {address}")
    
    # Clean the address
    clean_address = address.replace(',', '').replace('-', '')
    center_name = center_name.strip().replace(',', '').replace('-', '')
    
    # Prepare the API request
    params = {
        'address': center_name+clean_address,
        'language': 'English',
        'api_key': API_KEY
    }
    
    try:
        # Make the API request with a timeout of 15 seconds
        response = requests.get(BASE_URL, params=params, timeout=15)
        data = response.json()
        
        # Extract latitude and longitude from the response
        if data['status'] == 'ok' and data['geocodingResults']:
            location = data['geocodingResults'][0]['geometry']['location']
            latitude = location['lat']
            longitude = location['lng']
            print(f"Found location: {latitude}, {longitude}")
        else:
            latitude = longitude = None
            print(f"No location found for: {address}")
    
    except requests.exceptions.Timeout:
        print(f"Request timed out for address: {address}")
        latitude = longitude = None

    except requests.exceptions.RequestException as e:
        print(f"Request failed for address: {address}, error: {e}")
        latitude = longitude = None
    
    # Append the result to the new Excel sheet
    row_output = [sr_no, center_name, address, licence_no, latitude, longitude]
    sheet_output.append(row_output)

# Save the new Excel file
print("Saving output Excel file...")
wb_output.save(output_file)

print(f"Geocoding complete. Results saved to {output_file}")
